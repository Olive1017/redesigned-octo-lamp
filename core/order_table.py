"""单号对照表 - 从父文件夹内的 xlsx 读取 车号 → {交货单号, 订单号} 映射。

同一辆车当天只有一单，车号唯一，可直接作主键。
子文件夹名（人工按车牌命名）就是查询键。
"""

import os
import re
import glob
import logging
from typing import Optional, Dict
from openpyxl import load_workbook

log = logging.getLogger(__name__)

_列_车号 = "车号"
_列_交货单号 = "交货单号"
_列_订单号 = "订单号"


def 规整车牌(s: Optional[str]) -> str:
    """统一车牌/车号格式：去所有空白（含全角\u3000）、去分隔符、转大写。建表与查询两侧都要用它。"""
    if not s:
        return ""
    s = re.sub(r"\s+", "", str(s))
    s = s.replace(".", "").replace("-", "").replace("·", "")
    return s.upper()


class 单号表:
    def __init__(self, 映射: Dict[str, dict]):
        self._映射 = 映射

    def 查(self, 车牌: str) -> Optional[dict]:
        """返回 {"交货单号":..., "订单号":...}，查不到返回 None。"""
        return self._映射.get(规整车牌(车牌))

    def __len__(self):
        return len(self._映射)

    @classmethod
    def 从父文件夹(cls, 父文件夹: str) -> "单号表":
        路径 = _找唯一xlsx(父文件夹)
        wb = load_workbook(路径, read_only=True, data_only=True)
        ws = wb.active
        行s = list(ws.iter_rows(values_only=True))

        表头行号 = _定位表头(行s)
        if 表头行号 is None:
            raise RuntimeError(f"对照表未找到表头（需同时含「车号」「交货单号」）：{路径}")

        表头 = [(str(c).strip() if c is not None else "") for c in 行s[表头行号]]
        列 = {name: i for i, name in enumerate(表头)}
        for 必需 in (_列_车号, _列_交货单号, _列_订单号):
            if 必需 not in 列:
                raise RuntimeError(f"对照表缺少列「{必需}」：{路径}（现有列：{表头}）")

        映射: Dict[str, dict] = {}
        for 行 in 行s[表头行号 + 1:]:
            键 = 规整车牌(_取单元(行, 列[_列_车号]))
            if not 键:
                continue
            映射[键] = {
                "交货单号": _取文本(行, 列[_列_交货单号]),
                "订单号": _取文本(行, 列[_列_订单号]),
            }

        log.info("单号对照表已加载：%s，共 %d 条", os.path.basename(路径), len(映射))
        return cls(映射)


def _找唯一xlsx(父文件夹: str) -> str:
    候选 = [
        p for p in glob.glob(os.path.join(父文件夹, "*.xlsx"))
        if not os.path.basename(p).startswith("~$")  # 跳过 Excel 临时锁文件
    ]
    if not 候选:
        raise RuntimeError(f"父文件夹内未找到单号对照表 .xlsx：{父文件夹}")
    候选.sort()
    if len(候选) > 1:
        log.warning("父文件夹内有多张 xlsx，默认取第一张：%s", os.path.basename(候选[0]))
    return 候选[0]


def _定位表头(行s) -> Optional[int]:
    """扫描各行，找到同时含「车号」「交货单号」的那一行作为表头（容忍上方有标题行）。"""
    for i, 行 in enumerate(行s):
        文本集 = {(str(c).strip() if c is not None else "") for c in 行}
        if _列_车号 in 文本集 and _列_交货单号 in 文本集:
            return i
    return None


def _取单元(行, i: int):
    return 行[i] if i < len(行) else None


def _取文本(行, i: int) -> Optional[str]:
    """把单元格转成字符串；单号常被 Excel 存成数字，去掉 .0 尾巴。"""
    v = _取单元(行, i)
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None
